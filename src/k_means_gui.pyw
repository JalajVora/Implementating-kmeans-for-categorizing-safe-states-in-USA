import openpyxl
from random import choice
from Tkinter import *
import tkFileDialog

def saveToXl():
	# print "inside saveToXl"
	wb = openpyxl.load_workbook(fileName)
	ws = wb.create_sheet("result")
	row = 1
	for k in range(clusterNum) :
		ws.cell(row = row, column = 1).value = "Cluster "+str(k+1)
		ws.cell(row = row + 1, column = 1).value = "Centroid"
		for i in range(dimension) :
			ws.cell(row = row + 2, column = 1+i).value = clusterHeads[k][i]
		ws.cell(row = row + 3, column = 1).value = "Nodes"
		for i in range(len(clusters[k])) :
			for j in range(dimension) :
				ws.cell(row = row+4+i, column = 1+j).value = clusters[k][i][j]
		row += 5 + len(clusters[k])
	
	wb.save(fileName)

def chooseFile() :
	global fileName, fileNameAnswerLabel
	fileName = tkFileDialog.askopenfilename(filetypes = [("Excel File", ".xlsx")])
	# print fileName
	fileNameAnswerLabel.config(text = fileName)

def startClustering() :
	global num, clusterNum, dimension, clusters, clusterHeads, prevClusterHeads, nodes, root

	try:
		num = int(numEntry.get())
		dimension = int(dimensionEntry.get())
		clusterNum = int(clusterNumEntry.get())
	except Exception, e:
		# print str(e)
		return
	

	# print num, dimension, clusterNum

	wb = openpyxl.load_workbook(fileName)
	sh = wb.active


	#======================================
	#assigning values to nodes
	#======================================
	ws = list(sh.rows)
	for i in range(num) :
		temp = []
		for j in range(dimension) :
			temp.append(float(ws[i][j].value))
		nodes.append(tuple(temp))

	# for i in nodes :
	# 	print i

	#===============================
	#Choosing cluster heads randomly
	#===============================
	while len(clusterHeads) < clusterNum :
		a = choice(nodes)
		if a not in clusterHeads :
			clusterHeads.append(a)

	# print clusterHeads


	#===============================
	#Converge to clusters
	#===============================
	assignClusterHead()

	while prevClusterHeads != clusterHeads :
		assignClusterHead()

		prevClusterHeads = list(clusterHeads)

		# print "centroids"
		for i in range(clusterNum) :
			clusterHeads[i] = calcCentroid(clusters[i])

		# print clusterHeads
		# for i in clusterHeads :
		# 	print i

	# print clusters
	
	saveToXl()
	root.destroy()

def assignClusterHead() :
	global clusters
	clusters = []
	for i in range(clusterNum) :
		clusters.append([])

	for i in range(num) :
		temp = []
		# print nodes[i]
		for j in range(clusterNum) :
			d = calcDistance(nodes[i], clusterHeads[j])
			temp.append(d)
		# print temp
		c = temp.index(min(temp))
		clusters[c].append(nodes[i])

def calcCentroid(array) :
	temp = []
	for i in range(dimension) :
		tempSum = float(sum([j[i] for j in array]))/len(array)
		temp.append(tempSum)
	return tuple(temp)
	
def calcDistance(x, y) :
	dist = 0
	for i in range(dimension) :
		dist += (x[i]-y[i])**2
	dist = dist ** 0.5
	return dist


#======================================
#===============MAIN===================
#======================================

root = Tk()
root.title("k Means Executor")
num = 0
clusterNum = 0
dimension = 0
clusters = []
clusterHeads = []
prevClusterHeads = []
nodes = []
fileName = ""


#============ GUI stuff ===============


fileNameLabel = Label(root, text = "File Name : ")
fileNameLabel.grid(row = 0, column = 0, pady = 10)
browseButton = Button(root, text = "Browse", width = 10, command = chooseFile)
browseButton.grid(row = 0, column = 1, pady = 10)
fileNameAnswerLabel = Label(root, text = "")
fileNameAnswerLabel.grid(row = 0, column = 2, pady = 10, padx = 10)
numLabel = Label(root, text = "Number of entries : ")
numLabel.grid(row = 1, column = 0, pady = 10)
numEntry = Entry(root, width = 3)
numEntry.grid(row = 1, column = 1, pady = 10)
dimensionLabel = Label(root, text = "Number of Dimensions : ")
dimensionLabel.grid(row = 2, column = 0, pady = 10)
dimensionEntry = Entry(root, width = 3)
dimensionEntry.grid(row = 2, column = 1, pady = 10)
clusterNumLabel = Label(root, text = "Number of clusters : ")
clusterNumLabel.grid(row = 3, column = 0, pady = 10)
clusterNumEntry = Entry(root, width = 3)
clusterNumEntry.grid(row = 3, column = 1, pady = 10)
startButton = Button(root, text = "Start", width = 10, command = startClustering)
startButton.grid(row = 4, column = 0, pady = 10)
root.mainloop()




